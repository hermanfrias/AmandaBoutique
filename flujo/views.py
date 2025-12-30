
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from .models import MovimientoCaja, CotizacionDolar, ConfiguracionIVA
from .forms import MovimientoCajaForm, CotizacionDolarForm, ConfiguracionIVAForm

@login_required
def listar_movimientos(request):
    from django.utils.dateparse import parse_date
    
    movimientos = MovimientoCaja.objects.all()
    
    # Obtener filtros
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    tipo_filtro = request.GET.get('tipo')
    tipo_movimiento_filtro = request.GET.get('tipo_movimiento')
    metodo_pago_filtro = request.GET.get('metodo_pago')
    moneda_filtro = request.GET.get('moneda')
    
    # Aplicar filtro de fechas
    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None
    
    if fecha_inicio and fecha_fin:
        movimientos = movimientos.filter(fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        movimientos = movimientos.filter(fecha__gte=fecha_inicio)
    elif fecha_fin:
        movimientos = movimientos.filter(fecha__lte=fecha_fin)
    
    # Aplicar filtro de tipo
    if tipo_filtro and tipo_filtro != 'Todos':
        movimientos = movimientos.filter(tipo=tipo_filtro)
    
    # Aplicar filtro de tipo_movimiento
    if tipo_movimiento_filtro and tipo_movimiento_filtro != 'Todos':
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento_filtro)
    
    # Aplicar filtro de metodo_pago
    if metodo_pago_filtro and metodo_pago_filtro != 'Todos':
        movimientos = movimientos.filter(metodo_pago=metodo_pago_filtro)
    
    # Aplicar filtro de moneda
    if moneda_filtro and moneda_filtro != 'Todos':
        movimientos = movimientos.filter(moneda=moneda_filtro)
    
    # Calcular totales separados por moneda
    from django.db.models import Sum, Q
    
    # Total en Bolívares (solo movimientos en Bs)
    total_bs = movimientos.filter(moneda='Bs').aggregate(total=Sum('monto'))['total'] or 0
    
    # Total en Dólares (solo movimientos en $)
    total_usd = movimientos.filter(moneda='$').aggregate(total=Sum('monto'))['total'] or 0
    
    # Total en USD (todos los movimientos convertidos)
    total_usd_converted = movimientos.aggregate(total=Sum('monto_usd'))['total'] or 0
    
    context = {
        'movimientos': movimientos,
        'fecha_inicio': fecha_inicio_str or '',
        'fecha_fin': fecha_fin_str or '',
        'tipo_filtro': tipo_filtro or 'Todos',
        'tipo_movimiento_filtro': tipo_movimiento_filtro or 'Todos',
        'metodo_pago_filtro': metodo_pago_filtro or 'Todos',
        'moneda_filtro': moneda_filtro or 'Todos',
        'total_bs': round(total_bs, 2),
        'total_usd': round(total_usd, 2),
        'total_usd_converted': round(total_usd_converted, 2),
    }
    
    return render(request, 'flujo/listar_movimientos.html', context)


@login_required
@permission_required('flujo.add_movimientocaja', raise_exception=True)
def crear_movimiento(request):
    if request.method=='POST':
        form = MovimientoCajaForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request,'Movimiento guardado correctamente')
                return redirect('listar_movimientos')
            except ValidationError as e:
                form.add_error('fecha', e)
    else:
        form = MovimientoCajaForm()
    return render(request,'flujo/crear_movimiento.html',{'form':form})

@login_required
@permission_required('flujo.change_movimientocaja', raise_exception=True)
def editar_movimiento(request, id):
    movimiento = MovimientoCaja.objects.get(id=id)
    if request.method == 'POST':
        form = MovimientoCajaForm(request.POST, instance=movimiento)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Movimiento actualizado correctamente')
                return redirect('listar_movimientos')
            except ValidationError as e:
                form.add_error('fecha', e)
    else:
        form = MovimientoCajaForm(instance=movimiento)
    return render(request, 'flujo/editar_movimiento.html', {'form': form, 'movimiento': movimiento})

@login_required
def ver_movimiento(request, id):
    movimiento = MovimientoCaja.objects.get(id=id)
    return render(request, 'flujo/ver_movimiento.html', {'movimiento': movimiento})

@login_required
@permission_required('flujo.delete_movimientocaja', raise_exception=True)
def eliminar_movimiento(request, id):
    movimiento = MovimientoCaja.objects.get(id=id)
    if request.method == 'POST':
        movimiento.delete()
        messages.success(request, 'Movimiento eliminado correctamente')
        return redirect('listar_movimientos')
    return render(request, 'flujo/eliminar_movimiento.html', {'movimiento': movimiento})

@login_required
def listar_cotizaciones(request):
    cotizaciones = CotizacionDolar.objects.all()
    return render(request,'flujo/listar_cotizaciones.html',{'cotizaciones':cotizaciones})

@login_required
@permission_required('flujo.add_cotizaciondolar', raise_exception=True)
def crear_cotizacion(request):
    if request.method=='POST':
        form = CotizacionDolarForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,'Cotización guardada correctamente')
            return redirect('listar_cotizaciones')
    else:
        form = CotizacionDolarForm()
    return render(request,'flujo/crear_cotizacion.html',{'form':form})

@login_required
@permission_required('flujo.change_cotizaciondolar', raise_exception=True)
def editar_cotizacion(request, id):
    cotizacion = CotizacionDolar.objects.get(id=id)
    if request.method == 'POST':
        form = CotizacionDolarForm(request.POST, instance=cotizacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cotización actualizada correctamente')
            return redirect('listar_cotizaciones')
    else:
        form = CotizacionDolarForm(instance=cotizacion)
    return render(request, 'flujo/editar_cotizacion.html', {'form': form, 'cotizacion': cotizacion})

@login_required
@permission_required('flujo.delete_cotizaciondolar', raise_exception=True)
def eliminar_cotizacion(request, id):
    cotizacion = CotizacionDolar.objects.get(id=id)
    if request.method == 'POST':
        cotizacion.delete()
        messages.success(request, 'Cotización eliminada correctamente')
        return redirect('listar_cotizaciones')
    return render(request, 'flujo/eliminar_cotizacion.html', {'cotizacion': cotizacion})


# ==================== VISTAS PARA CONFIGURACION IVA ====================

@login_required
def listar_configuraciones_iva(request):
    configuraciones = ConfiguracionIVA.objects.all()
    return render(request, 'flujo/listar_configuraciones_iva.html', {'configuraciones': configuraciones})

@login_required
@permission_required('flujo.add_configuracioniva', raise_exception=True)
def crear_configuracion_iva(request):
    if request.method == 'POST':
        form = ConfiguracionIVAForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración de IVA guardada correctamente')
            return redirect('listar_configuraciones_iva')
    else:
        form = ConfiguracionIVAForm()
    return render(request, 'flujo/crear_configuracion_iva.html', {'form': form})

@login_required
@permission_required('flujo.change_configuracioniva', raise_exception=True)
def editar_configuracion_iva(request, id):
    configuracion = ConfiguracionIVA.objects.get(id=id)
    if request.method == 'POST':
        form = ConfiguracionIVAForm(request.POST, instance=configuracion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración de IVA actualizada correctamente')
            return redirect('listar_configuraciones_iva')
    else:
        form = ConfiguracionIVAForm(instance=configuracion)
    return render(request, 'flujo/editar_configuracion_iva.html', {'form': form, 'configuracion': configuracion})

@login_required
@permission_required('flujo.delete_configuracioniva', raise_exception=True)
def eliminar_configuracion_iva(request, id):
    configuracion = ConfiguracionIVA.objects.get(id=id)
    if request.method == 'POST':
        configuracion.delete()
        messages.success(request, 'Configuración de IVA eliminada correctamente')
        return redirect('listar_configuraciones_iva')
    return render(request, 'flujo/eliminar_configuracion_iva.html', {'configuracion': configuracion})

@login_required
def dashboard_flujo(request):
    import datetime
    import json
    import traceback
    
    try:
        # Obtener filtros
        mes_actual = datetime.date.today().month
        anio_actual = datetime.date.today().year
        
        # Verificar si hay años disponibles
        anios_disponibles = MovimientoCaja.objects.dates('fecha', 'year')
        if not anios_disponibles:
            # Si no hay movimientos, usar año actual
            anios_disponibles = [datetime.date.today()]
        
        mes_filtro = request.GET.get('mes', str(mes_actual))
        anio_filtro = request.GET.get('anio', str(anio_actual))
        
        # Limpiar el año (remover separadores de miles como 2.025 -> 2025)
        anio_filtro = anio_filtro.replace('.', '').replace(',', '')
        
        movimientos = MovimientoCaja.objects.all()
        
        # Filtrar por mes y año si se seleccionan
        if mes_filtro and anio_filtro:
            movimientos = movimientos.filter(fecha__month=mes_filtro, fecha__year=anio_filtro)
        
        # Totales USD (Mes)
        total_ingresos_usd = movimientos.filter(tipo='Ingreso').aggregate(total=Sum('monto_usd'))['total'] or 0
        total_gastos_usd = movimientos.filter(tipo='Egresos').aggregate(total=Sum('monto_usd'))['total'] or 0
        saldo_usd = total_ingresos_usd - total_gastos_usd
        
        # Cantidades (Mes)
        cant_ingresos = movimientos.filter(tipo='Ingreso').count()
        cant_gastos = movimientos.filter(tipo='Egresos').count()
        
        # Rentabilidad (Mes)
        rentabilidad = 0
        if total_ingresos_usd > 0:
            rentabilidad = (saldo_usd / total_ingresos_usd) * 100

        # Totales Bs (Estimado - Mes)
        total_ingresos_bs = 0
        total_gastos_bs = 0
        for m in movimientos:
            if m.moneda == 'Bs':
                factor = 1
            else:
                try:
                    cot = CotizacionDolar.objects.get(fecha=m.fecha)
                    factor = cot.valor
                except CotizacionDolar.DoesNotExist:
                    factor = 1
            if m.tipo == 'Ingreso':
                total_ingresos_bs += m.monto_usd * factor
            else:
                total_gastos_bs += m.monto_usd * factor
        saldo_bs = total_ingresos_bs - total_gastos_bs

        # --- RESUMEN ANUAL ---
        movimientos_anual = MovimientoCaja.objects.filter(fecha__year=anio_filtro)
        total_ingresos_usd_anual = movimientos_anual.filter(tipo='Ingreso').aggregate(total=Sum('monto_usd'))['total'] or 0
        total_gastos_usd_anual = movimientos_anual.filter(tipo='Egresos').aggregate(total=Sum('monto_usd'))['total'] or 0
        saldo_usd_anual = total_ingresos_usd_anual - total_gastos_usd_anual
        
        rentabilidad_anual = 0
        if total_ingresos_usd_anual > 0:
            rentabilidad_anual = (saldo_usd_anual / total_ingresos_usd_anual) * 100

        # --- RESUMEN ACUMULATIVO TOTAL ---
        movimientos_total = MovimientoCaja.objects.all()
        total_ingresos_usd_total = movimientos_total.filter(tipo='Ingreso').aggregate(total=Sum('monto_usd'))['total'] or 0
        total_gastos_usd_total = movimientos_total.filter(tipo='Egresos').aggregate(total=Sum('monto_usd'))['total'] or 0
        saldo_usd_total = total_ingresos_usd_total - total_gastos_usd_total
        
        rentabilidad_total = 0
        if total_ingresos_usd_total > 0:
            rentabilidad_total = (saldo_usd_total / total_ingresos_usd_total) * 100

        # Gráfico (Mantenemos la lógica global para el gráfico o la ajustamos al año seleccionado?)
        # Para el gráfico es mejor mostrar el año seleccionado completo
        movimientos_anio = MovimientoCaja.objects.filter(fecha__year=anio_filtro)
        
        meses=[]
        ingresos_mes=[]
        gastos_mes=[]
        queryset = movimientos_anio.annotate(mes=TruncMonth('fecha')).values('mes').distinct().order_by('mes')

        for q in queryset:
            mes=q['mes'].strftime('%B')
            meses.append(mes)
            ingresos_mes.append(round(movimientos_anio.filter(tipo='Ingreso',fecha__month=q['mes'].month).aggregate(total=Sum('monto_usd'))['total'] or 0,2))
            gastos_mes.append(round(movimientos_anio.filter(tipo='Egresos',fecha__month=q['mes'].month).aggregate(total=Sum('monto_usd'))['total'] or 0,2))

        # Serializar para JavaScript
        meses_json = json.dumps(meses)
        ingresos_mes_json = json.dumps([float(x) for x in ingresos_mes])
        gastos_mes_json = json.dumps([float(x) for x in gastos_mes])
        
        context={
            'total_ingresos_usd': round(total_ingresos_usd,2),
            'total_gastos_usd': round(total_gastos_usd,2),
            'saldo_usd': round(saldo_usd,2),
            'rentabilidad': round(rentabilidad, 2),
            'cant_ingresos': cant_ingresos,
            'cant_gastos': cant_gastos,
            'total_ingresos_bs': round(total_ingresos_bs,2),
            'total_gastos_bs': round(total_gastos_bs,2),
            'saldo_bs': round(saldo_bs,2),
            'meses': meses_json,
            'ingresos_mes': ingresos_mes_json,
            'gastos_mes': gastos_mes_json,
            'mes_filtro': int(mes_filtro),
            'anio_filtro': int(anio_filtro),
            'anios_disponibles': anios_disponibles,
            # Anual
            'total_ingresos_usd_anual': round(total_ingresos_usd_anual, 2),
            'total_gastos_usd_anual': round(total_gastos_usd_anual, 2),
            'saldo_usd_anual': round(saldo_usd_anual, 2),
            'rentabilidad_anual': round(rentabilidad_anual, 2),
            # Total
            'total_ingresos_usd_total': round(total_ingresos_usd_total, 2),
            'total_gastos_usd_total': round(total_gastos_usd_total, 2),
            'saldo_usd_total': round(saldo_usd_total, 2),
            'rentabilidad_total': round(rentabilidad_total, 2),
        }

        return render(request,'flujo/dashboard.html',context)
    except Exception as e:
        # Imprimir el error completo en la consola
        print("=" * 80)
        print("ERROR EN DASHBOARD:")
        print(traceback.format_exc())
        print("=" * 80)
        from django.http import HttpResponse
        return HttpResponse(f"Error: {str(e)}<br><br><pre>{traceback.format_exc()}</pre>", status=500)

@login_required
def movimientos_pdf(request):
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from weasyprint import HTML, CSS
    from django.conf import settings
    from django.utils.dateparse import parse_date
    from django.db.models import Q
    import os
    import traceback

    try:
        movimientos = MovimientoCaja.objects.all().order_by('-fecha')

        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        tipo_filtro = request.GET.get('tipo')
        tipo_movimiento_filtro = request.GET.get('tipo_movimiento')
        metodo_pago_filtro = request.GET.get('metodo_pago')
        moneda_filtro = request.GET.get('moneda')

        fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
        fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None

        if fecha_inicio and fecha_fin:
            movimientos = movimientos.filter(fecha__range=[fecha_inicio, fecha_fin])
        elif fecha_inicio:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio)
        elif fecha_fin:
            movimientos = movimientos.filter(fecha__lte=fecha_fin)
        
        # Aplicar filtro de tipo
        if tipo_filtro and tipo_filtro != 'Todos':
            movimientos = movimientos.filter(tipo=tipo_filtro)
        
        # Aplicar filtro de tipo_movimiento
        if tipo_movimiento_filtro and tipo_movimiento_filtro != 'Todos':
            movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento_filtro)
        
        # Aplicar filtro de metodo_pago
        if metodo_pago_filtro and metodo_pago_filtro != 'Todos':
            movimientos = movimientos.filter(metodo_pago=metodo_pago_filtro)
        
        # Aplicar filtro de moneda
        if moneda_filtro and moneda_filtro != 'Todos':
            movimientos = movimientos.filter(moneda=moneda_filtro)
        
        # Calcular totales por tipo (Ingreso/Gasto)
        total_ingresos_usd = movimientos.filter(tipo='Ingreso').aggregate(total=Sum('monto_usd'))['total'] or 0
        total_gastos_usd = movimientos.filter(tipo='Egresos').aggregate(total=Sum('monto_usd'))['total'] or 0
        saldo_usd = total_ingresos_usd - total_gastos_usd
        
        # Calcular totales separados por moneda
        total_bs = movimientos.filter(moneda='Bs').aggregate(total=Sum('monto'))['total'] or 0
        total_usd = movimientos.filter(moneda='$').aggregate(total=Sum('monto'))['total'] or 0
        total_usd_converted = movimientos.aggregate(total=Sum('monto_usd'))['total'] or 0

        html_string = render_to_string('flujo/movimientos_pdf.html', {
            'movimientos': movimientos,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'tipo_filtro': tipo_filtro or '',
            'tipo_movimiento_filtro': tipo_movimiento_filtro or '',
            'metodo_pago_filtro': metodo_pago_filtro or '',
            'moneda_filtro': moneda_filtro or '',
            'total_ingresos_usd': round(total_ingresos_usd, 2),
            'total_gastos_usd': round(total_gastos_usd, 2),
            'saldo_usd': round(saldo_usd, 2),
            'total_bs': round(total_bs, 2),
            'total_usd': round(total_usd, 2),
            'total_usd_converted': round(total_usd_converted, 2),
        })


        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="movimientos_caja.pdf"'

        css_path = os.path.join(settings.STATICFILES_DIRS[0], "BoutiqueApp/css/pdf.css")

        HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
            response, stylesheets=[CSS(css_path)]
        )
        return response
    except Exception as e:
        # Mostrar el error completo para debugging
        error_msg = f"Error generando PDF: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print("=" * 80)
        print(error_msg)
        print("=" * 80)
        return HttpResponse(f"<pre>{error_msg}</pre>", status=500)

@login_required
def movimientos_excel(request):
    from django.http import HttpResponse
    from django.utils.dateparse import parse_date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    movimientos = MovimientoCaja.objects.all().order_by('-fecha')

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    tipo_filtro = request.GET.get('tipo')
    tipo_movimiento_filtro = request.GET.get('tipo_movimiento')
    metodo_pago_filtro = request.GET.get('metodo_pago')
    moneda_filtro = request.GET.get('moneda')

    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None

    if fecha_inicio and fecha_fin:
        movimientos = movimientos.filter(fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        movimientos = movimientos.filter(fecha__gte=fecha_inicio)
    elif fecha_fin:
        movimientos = movimientos.filter(fecha__lte=fecha_fin)
    
    # Aplicar filtro de tipo
    if tipo_filtro and tipo_filtro != 'Todos':
        movimientos = movimientos.filter(tipo=tipo_filtro)
    
    # Aplicar filtro de tipo_movimiento
    if tipo_movimiento_filtro and tipo_movimiento_filtro != 'Todos':
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento_filtro)
    
    # Aplicar filtro de metodo_pago
    if metodo_pago_filtro and metodo_pago_filtro != 'Todos':
        movimientos = movimientos.filter(metodo_pago=metodo_pago_filtro)
    
    # Aplicar filtro de moneda
    if moneda_filtro and moneda_filtro != 'Todos':
        movimientos = movimientos.filter(moneda=moneda_filtro)

    # Calcular totales
    total_ingresos_usd = movimientos.filter(tipo='Ingreso').aggregate(total=Sum('monto_usd'))['total'] or 0
    total_gastos_usd = movimientos.filter(tipo='Egresos').aggregate(total=Sum('monto_usd'))['total'] or 0
    saldo_usd = total_ingresos_usd - total_gastos_usd

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos de Caja"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Amanda Mateo Boutique - Movimientos de Caja'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Rango de fechas
    if fecha_inicio or fecha_fin:
        ws.merge_cells('A2:H2')
        if fecha_inicio and fecha_fin:
            ws['A2'] = f'Del {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
        elif fecha_inicio:
            ws['A2'] = f'Desde {fecha_inicio.strftime("%d/%m/%Y")}'
        elif fecha_fin:
            ws['A2'] = f'Hasta {fecha_fin.strftime("%d/%m/%Y")}'
        ws['A2'].alignment = Alignment(horizontal="center")

    # Encabezados
    start_row = 4 if (fecha_inicio or fecha_fin) else 3
    headers = ['Fecha', 'Descripción', 'Tipo', 'Tipo Mov.', 'Método Pago', 'Moneda', 'Monto', 'Monto USD']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    row = start_row + 1
    for mov in movimientos:
        ws.cell(row=row, column=1, value=mov.fecha.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=mov.descripcion)
        ws.cell(row=row, column=3, value=mov.tipo)
        ws.cell(row=row, column=4, value=mov.tipo_movimiento or "-")
        ws.cell(row=row, column=5, value=mov.metodo_pago or "-")
        ws.cell(row=row, column=6, value=mov.moneda)
        ws.cell(row=row, column=7, value=float(mov.monto))
        ws.cell(row=row, column=8, value=float(mov.monto_usd))
        row += 1

    # Totales
    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Ingresos USD:")
    ws.cell(row=row, column=2, value=f"${round(total_ingresos_usd, 2)}")
    row += 1
    ws.cell(row=row, column=1, value="Total Egresos USD:")
    ws.cell(row=row, column=2, value=f"${round(total_gastos_usd, 2)}")
    row += 1
    ws.cell(row=row, column=1, value="Saldo USD:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"${round(saldo_usd, 2)}").font = Font(bold=True)

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="movimientos_caja.xlsx"'
    wb.save(response)
    return response
